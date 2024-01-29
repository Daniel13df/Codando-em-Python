from openpyxl import load_workbook, workbook
import pyautogui
import os
import time
import sys

# Caixa de confirmação de continuidade do procedimento

def mensagem_confirmacao():
    msg_confirmacao = pyautogui.confirm(text="Deseja continuar?", title="Confirmação de continuação", buttons=["Sim", "Não"])
    if msg_confirmacao == "Não":
        sys.exit()


# -----------------------------------------------------


# Primeira etapa

caminho_script = os.path.dirname(os.path.abspath(__file__))

#Pega os dados de origem
dados = load_workbook(os.path.join(caminho_script,'seu/arquivo/de/dados'))
massa = load_workbook(os.path.join(caminho_script,'seu/arquivo/de/alteracao'))

report = dados['Report']
cells = report['A2': 'A31']

codigos = massa['codigos']
cells_massa = codigos['C3':'C32']

for i in range(len(cells)):
    cells_massa[i][0].value = cells[i][0].value

massa.save(os.path.join(caminho_script,'seu/arquivo/de/alteracao'))

pyautogui.alert("ID's Salvos com sucesso", "Automação Daniel Coelho")
mensagem_confirmacao()

# Segunda etapa

#Seleciona os dados da planilha de alteração
value_range = codigos['A3':'F32']

#Clica no campo em branco da API
pyautogui.click(x=-1724, y=629)
pyautogui.hotkey("Ctrl","a")
time.sleep(0.5)

#Acrescenta a informaão dentro da API
pyautogui.typewrite("mutation{\n")

#Faz uma repetição dos dados que que estão selecionados na planilha de alteração para jogar na API WEB
for a,b,c,d,e,f in value_range:
    time.sleep(0.5)
    pyautogui.typewrite(f'{a.value}{b.value}{c.value}{d.value}{e.value}{f.value}\n')

time.sleep(0.5)
pyautogui.click(x=-2162, y=156)
